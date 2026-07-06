"""
tracker/views_reports_matrix.py

Matrix (pivot) reporting — the "Wayne grid." ONE endpoint reproduces every one
of the client's spreadsheet tabs by swapping the ROW axis; CLIENT is always the
column axis (his layout is fixed: clients across the top, a Total column on the
right).

    rows=employee  → JUN2026 tab   (person × client)
    rows=day       → work-effort    (day-of-month × client)
    rows=month     → Summary tab    (month × client, time-series)

WHY A SEPARATE ENDPOINT INSTEAD OF WIDENING reports_summary:
  reports_summary answers "totals per employee OR per client." A matrix needs
  the full (row_axis × client) cross. Rather than write a second block-walker
  (which would inevitably drift from _aggregate's billable/material/anomaly
  rules and re-introduce the today_time-vs-block mismatch), this REUSES the
  exact same _block_queryset + _block_minutes + _is_billable_block + materiality
  helpers from views_reports. The cell value is computed by the identical rules
  _aggregate uses, so a matrix row total equals that employee's reports_summary
  total to the minute. Single source of truth preserved.

SCALABILITY:
  - The grid is dense but computed in ONE pass over the same block list.
  - Column set = clients that actually have time in the window (not every
    client the firm has ever had), so a 500-client firm only renders the
    columns with data. A full-catalog option can be added later via a flag.
  - "Expand all / no pagination" is a PRESENTATION concern (OrgReportPreset +
    frontend virtualization), not a data concern — the endpoint always returns
    the full grid; the client decides whether to paginate.

Wire into urls.py:
    path("api/reports/matrix/",        views_reports_matrix.reports_matrix,        name="reports_matrix"),
    path("api/reports/matrix/export/", views_reports_matrix.reports_matrix_export, name="reports_matrix_export"),
    path("api/reports/presets/",       views_reports_matrix.reports_presets,       name="reports_presets"),
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import timedelta

from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# Reuse EVERYTHING from the existing report module — same counting rules, same
# scope/impersonation resolution. This is the whole point: no second source of
# truth for what a block's minutes are or whether it's billable.
from tracker.views_reports import (
    _block_queryset,
    _block_minutes,
    _is_billable_block,
    _is_material,
    _dominant_category,
    _resolve_scope,
    _resolve_period_window,
    _day_bounds_utc,
    _EXCLUDE_CATEGORIES,
    _TRUNC,
)
from tracker.views import get_request_org_override


# Which metric each cell carries. Mirrors the three numbers reports_summary
# already exposes per row, so a matrix in "billable" mode sums to the same
# billable total reports_summary shows.
_VALID_METRICS = {"total", "billable", "non_billable"}

# Row axes. "employee" and the time buckets are the only ones that make sense as
# ROWS given CLIENT is the fixed column axis. (client×client would be nonsense.)
_VALID_ROW_AXES = {"employee", "day", "week", "month", "quarter", "year"}


def _row_key_and_label(b, row_axis: str):
    """
    Return (key, label, sort_key) for a block under the chosen row axis.

    employee → (user_id, display name, name)
    day/…    → (bucket ISO date, human label, ISO date)  — local-time bucketed
    """
    if row_axis == "employee":
        key = b.user_id
        label = (
            (b.user.get_full_name().strip() or b.user.username)
            if b.user_id else "Unknown"
        )
        return key, label, (label or "").lower()

    # time bucket — localize then truncate to the axis granularity
    local = timezone.localtime(b.start)
    d = local.date()
    if row_axis == "day":
        key = d.isoformat()
        # Wayne's work-effort sheet labels rows by day-of-month number.
        label = str(d.day)
        return key, label, key
    if row_axis == "week":
        monday = d - timedelta(days=d.weekday())
        return monday.isoformat(), monday.isoformat(), monday.isoformat()
    if row_axis == "month":
        first = d.replace(day=1)
        return first.isoformat(), first.strftime("%b %Y"), first.isoformat()
    if row_axis == "quarter":
        qsm = ((d.month - 1) // 3) * 3 + 1
        first = d.replace(month=qsm, day=1)
        q = (d.month - 1) // 3 + 1
        return first.isoformat(), f"Q{q} {d.year}", first.isoformat()
    if row_axis == "year":
        first = d.replace(month=1, day=1)
        return first.isoformat(), str(d.year), first.isoformat()
    # fallback
    return b.user_id, "Unknown", ""


def _cell_minutes(b, metric: str) -> int:
    """
    Minutes this block contributes to a cell under `metric`, applying the SAME
    rules _aggregate uses so matrix totals reconcile with reports_summary:

      - anomalous blocks → 0 (via _block_minutes)
      - genuine idle/uncategorized noise with no client → dropped
      - non-billable sub-2min slivers → dropped (immaterial), EXCEPT billable-
        with-client time always counts (matches _aggregate's fix)
    """
    minutes = _block_minutes(b)
    if minutes <= 0:
        return 0

    cat = _dominant_category(b).lower()
    billable = _is_billable_block(b)

    # Drop non-billable idle/uncategorized noise (mirror _aggregate).
    if cat in _EXCLUDE_CATEGORIES and not billable:
        return 0
    # Immaterial non-billable slivers are noise; billable-with-client always counts.
    if not _is_material(b) and not billable:
        return 0

    if metric == "billable":
        return minutes if billable else 0
    if metric == "non_billable":
        return 0 if billable else minutes
    return minutes  # "total"


def _build_matrix(blocks, row_axis: str, metric: str):
    """
    One pass → dense grid.

    Returns:
      columns: [{id, label}]           # clients with data, biggest first
      rows:    [{key, label, cells:{col_id: hours}, total_hours}]
      col_totals: {col_id: hours}
      grand_total_hours
    """
    # (row_key) -> {"label","sort","cells": {client_key: minutes}, "total": min}
    rows: dict = defaultdict(lambda: {
        "label": "", "sort": "", "cells": defaultdict(int), "total": 0,
    })
    # client_key -> {"label", "total_min"}
    cols: dict = defaultdict(lambda: {"label": "", "total": 0})
    grand = 0

    for b in blocks:
        minutes = _cell_minutes(b, metric)
        if minutes <= 0:
            continue

        rkey, rlabel, rsort = _row_key_and_label(b, row_axis)
        # Client is the column. No-client time folds into an explicit bucket so
        # the grid sums to the same grand total (rather than silently dropping).
        ckey = b.client_id if b.client_id else "__none__"
        clabel = b.client.name if b.client_id else "No Client"

        r = rows[rkey]
        r["label"] = rlabel
        r["sort"] = rsort
        r["cells"][ckey] += minutes
        r["total"] += minutes

        c = cols[ckey]
        c["label"] = clabel
        c["total"] += minutes

        grand += minutes

    # Column ordering: biggest first (matches the rest of the product's "lead
    # with the busiest" convention). "No Client" always sinks to the end.
    col_items = sorted(
        cols.items(),
        key=lambda kv: (kv[0] == "__none__", -kv[1]["total"]),
    )
    columns = [{"id": (None if k == "__none__" else k), "key": k,
                "label": v["label"], "total_hours": round(v["total"] / 60, 2)}
               for k, v in col_items]

    # Row ordering:
    #   employee → biggest total first (owner scans top performers)
    #   time axis → chronological (a day/month grid must read in order)
    if row_axis == "employee":
        row_items = sorted(rows.items(), key=lambda kv: -kv[1]["total"])
    else:
        row_items = sorted(rows.items(), key=lambda kv: kv[1]["sort"])

    out_rows = []
    for rkey, r in row_items:
        cells = {}
        for c in columns:
            m = r["cells"].get(c["key"], 0)
            cells[str(c["key"])] = round(m / 60, 2) if m else 0.0
        out_rows.append({
            "key": (None if rkey == "__none__" else rkey),
            "label": r["label"],
            "cells": cells,
            "total_hours": round(r["total"] / 60, 2),
        })

    return {
        "columns": columns,
        "rows": out_rows,
        "grand_total_hours": round(grand / 60, 2),
    }


def _resolve_window(request):
    """Shared window resolution — custom start/end overrides named period."""
    period = (request.GET.get("period") or "month").lower()
    if period not in _TRUNC:
        period = "month"
    date_str = request.GET.get("date")
    anchor = parse_date(date_str) if date_str else timezone.localdate()
    if not anchor:
        anchor = timezone.localdate()

    custom_start = parse_date(request.GET.get("start") or "")
    custom_end = parse_date(request.GET.get("end") or "")
    if custom_start and custom_end and custom_start <= custom_end:
        return period, custom_start, custom_end
    start_date, end_date = _resolve_period_window(period, anchor)
    return period, start_date, end_date


# ──────────────────────────────────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────────────────────────────────
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def reports_matrix(request):
    """
    GET /api/reports/matrix/
      ?rows=employee|day|week|month|quarter|year   (default: employee)
      &metric=total|billable|non_billable           (default: billable)
      &period=day|week|month|quarter|year           (default: month)
      &date=YYYY-MM-DD                               (anchor)
      &start=YYYY-MM-DD&end=YYYY-MM-DD               (custom range; overrides period)
      &org_id=<id>                                   (staff impersonation)

    Client is ALWAYS the column axis. Returns a dense grid + row/col totals.
    Role-gating identical to reports_summary (members locked to self).
    """
    row_axis = (request.GET.get("rows") or "employee").lower()
    if row_axis not in _VALID_ROW_AXES:
        row_axis = "employee"
    metric = (request.GET.get("metric") or "billable").lower()
    if metric not in _VALID_METRICS:
        metric = "billable"

    org = get_request_org_override(request)
    if not org:
        return Response({"error": "No organization found"}, status=404)

    can_see_all, forced_user_id = _resolve_scope(request, org)
    period, start_date, end_date = _resolve_window(request)
    start_utc, end_utc = _day_bounds_utc(start_date, end_date)

    blocks = list(
        _block_queryset(org, start_utc, end_utc, can_see_all, forced_user_id)
    )
    matrix = _build_matrix(blocks, row_axis, metric)

    return Response({
        "org_id": org.id,
        "org_name": org.name,
        "rows_axis": row_axis,
        "cols_axis": "client",
        "metric": metric,
        "period": period,
        "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
        "scope": "all" if can_see_all else "self",
        "columns": matrix["columns"],
        "rows": matrix["rows"],
        "grand_total_hours": matrix["grand_total_hours"],
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def reports_matrix_export(request):
    """
    GET /api/reports/matrix/export/  — same params as reports_matrix.
    Writes the client-columns grid to XLSX in Wayne's layout: row-axis label in
    col A, one column per client, Total column on the right, a totals row at the
    bottom. CPAs live in Excel; this is the deliverable he actually wanted.
    """
    # Lazy import so the module loads even if openpyxl isn't present at import.
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    row_axis = (request.GET.get("rows") or "employee").lower()
    if row_axis not in _VALID_ROW_AXES:
        row_axis = "employee"
    metric = (request.GET.get("metric") or "billable").lower()
    if metric not in _VALID_METRICS:
        metric = "billable"

    org = get_request_org_override(request)
    if not org:
        return Response({"error": "No organization found"}, status=404)

    can_see_all, forced_user_id = _resolve_scope(request, org)
    period, start_date, end_date = _resolve_window(request)
    start_utc, end_utc = _day_bounds_utc(start_date, end_date)

    blocks = list(
        _block_queryset(org, start_utc, end_utc, can_see_all, forced_user_id)
    )
    matrix = _build_matrix(blocks, row_axis, metric)

    # ── Palette — one emerald family + neutral slate, matching the web app ──
    INK       = "0F172A"   # slate-900 text
    MUTED     = "64748B"   # slate-500 subtitle
    HEADER_BG = "0F3D2E"   # deep emerald header band
    HEADER_FG = "FFFFFF"
    TOTALCOL  = "064E3B"   # emerald-900 for the Total column header
    BAND_BG   = "F1F5F9"   # slate-100 zebra band
    NOCLIENT  = "FFF7ED"   # amber-50 tint marking the "No Client" catch-all column
    TOTCOL_BG = "ECFDF5"   # faint emerald tint down the Total column
    TOTROW_BG = "DCFCE7"   # emerald-100 totals band
    GRAND_BG  = "16A34A"   # emerald-600 grand-total cell
    GRID      = "E2E8F0"   # slate-200 gridlines

    metric_label = {
        "billable": "Billable hours",
        "non_billable": "Non-billable hours",
        "total": "Total hours",
    }.get(metric, metric)
    axis_label = {
        "employee": "Employee", "day": "Day of Month", "week": "Week",
        "month": "Month", "quarter": "Quarter", "year": "Year",
    }[row_axis]

    wb = Workbook()
    ws = wb.active
    ws.title = metric_label[:31]          # Excel caps sheet titles at 31 chars
    ws.sheet_view.showGridLines = False   # our own borders read cleaner

    thin = Side(style="thin", color=GRID)
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = matrix["columns"]
    ncols = 1 + len(cols) + 1              # axis + clients + Total
    last_col_letter = get_column_letter(ncols)
    # Which physical column (if any) is the "No Client" bucket, so we can tint it.
    noclient_idx = next(
        (j for j, c in enumerate(cols, start=2) if c["key"] == "__none__"), None
    )

    # ── Title band (row 1) + subtitle (row 2) ──────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=org.name)
    t.font = Font(bold=True, size=15, color=INK)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(
        row=2, column=1,
        value=f"{axis_label} × Client  ·  {metric_label}  ·  {start_date} to {end_date}",
    )
    sub.font = Font(size=10, color=MUTED, italic=True)
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Header row (row 4) ─────────────────────────────────────────────
    hr = 4
    ws.row_dimensions[hr].height = 30

    def _style_header(cell, bg=HEADER_BG, align="center"):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=HEADER_FG, size=10)
        cell.border = grid_border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    _style_header(ws.cell(row=hr, column=1, value=axis_label), align="left")
    for j, col in enumerate(cols, start=2):
        _style_header(ws.cell(row=hr, column=j, value=col["label"]))
    _style_header(ws.cell(row=hr, column=ncols, value="Total"), bg=TOTALCOL)

    # ── Data rows — zebra banding, blank on zero (mirrors the on-screen "·") ──
    NUM_FMT = "#,##0.00;-#,##0.00;\"\""
    r = hr + 1
    for i, row in enumerate(matrix["rows"]):
        band = PatternFill("solid", fgColor=BAND_BG) if i % 2 else None

        lc = ws.cell(row=r, column=1, value=row["label"])
        lc.font = Font(bold=True, color=INK, size=10)
        lc.border = grid_border
        lc.alignment = Alignment(horizontal="left", vertical="center")
        if band:
            lc.fill = band

        for j, col in enumerate(cols, start=2):
            v = row["cells"].get(str(col["key"]), 0) or 0
            cell = ws.cell(row=r, column=j, value=round(v, 2))
            cell.border = grid_border
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if j == noclient_idx:
                cell.fill = PatternFill("solid", fgColor=NOCLIENT)
            elif band:
                cell.fill = band

        tcell = ws.cell(row=r, column=ncols, value=round(row["total_hours"], 2))
        tcell.font = Font(bold=True, color=INK, size=10)
        tcell.border = grid_border
        tcell.number_format = NUM_FMT
        tcell.alignment = Alignment(horizontal="right", vertical="center")
        tcell.fill = PatternFill("solid", fgColor=TOTCOL_BG)
        r += 1

    # ── Totals band ────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 22
    tot = ws.cell(row=r, column=1, value="Total")
    tot.font = Font(bold=True, color=INK, size=10)
    tot.fill = PatternFill("solid", fgColor=TOTROW_BG)
    tot.border = grid_border
    tot.alignment = Alignment(horizontal="left", vertical="center")
    for j, col in enumerate(cols, start=2):
        cell = ws.cell(row=r, column=j, value=round(col["total_hours"], 2))
        cell.font = Font(bold=True, color=INK, size=10)
        cell.fill = PatternFill("solid", fgColor=TOTROW_BG)
        cell.border = grid_border
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal="right", vertical="center")
    g = ws.cell(row=r, column=ncols, value=round(matrix["grand_total_hours"], 2))
    g.font = Font(bold=True, color="FFFFFF", size=11)
    g.fill = PatternFill("solid", fgColor=GRAND_BG)
    g.border = grid_border
    g.number_format = "#,##0.00"
    g.alignment = Alignment(horizontal="right", vertical="center")

    # ── Column widths (autofit-ish from label length) + freeze panes ───
    ws.column_dimensions["A"].width = max(16, min(32, max(
        (len(str(row["label"])) for row in matrix["rows"]), default=10) + 3))
    for j, col in enumerate(cols, start=2):
        ws.column_dimensions[get_column_letter(j)].width = max(
            10, min(22, len(str(col["label"])) + 2))
    ws.column_dimensions[last_col_letter].width = 11
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)  # freeze title/header + axis col
    ws.print_title_rows = f"{hr}:{hr}"               # repeat header when printed

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    slug = getattr(org, "slug", None) or org.id
    fname = f"matrix_{slug}_{row_axis}_x_client_{metric}_{start_date}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ──────────────────────────────────────────────────────────────────────────
# Report presets — "custom for TL Wall, scalable for everyone"
# ──────────────────────────────────────────────────────────────────────────
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def reports_presets(request):
    """
    GET  /api/reports/presets/?org_id=   → list this org's saved layouts.
    POST /api/reports/presets/           → create/update a layout (owners/staff).

    A preset is just a saved (rows, metric, expand_all, ...) layout. TL Wall
    gets a preset with expand_all=True, paginate=False. Every other firm keeps
    the default. The "custom" behavior is DATA, not a code fork — which is how
    this stays scalable.

    Requires the OrgReportPreset model (see model stub below). Import is lazy so
    this module still loads before the migration is applied.
    """
    from tracker.models import OrgReportPreset  # lazy (post-migration)

    org = get_request_org_override(request)
    if not org:
        return Response({"error": "No organization found"}, status=404)

    if request.method == "GET":
        out = []
        for p in OrgReportPreset.objects.filter(org=org).order_by("-is_default", "name"):
            out.append({
                "id": p.id,
                "name": p.name,
                "rows_axis": p.rows_axis,
                "metric": p.metric,
                "expand_all": p.expand_all,
                "paginate": p.paginate,
                "show_totals": p.show_totals,
                "is_default": p.is_default,
            })
        return Response({"org_id": org.id, "presets": out})

    # POST — only owners/admins/managers or staff may define firm layouts.
    can_see_all, _ = _resolve_scope(request, org)
    if not can_see_all:
        return Response({"error": "Only managers can save report layouts."}, status=403)

    data = request.data or {}
    name = (data.get("name") or "").strip()
    if not name:
        return Response({"error": "name is required."}, status=400)

    rows_axis = (data.get("rows_axis") or "employee").lower()
    if rows_axis not in _VALID_ROW_AXES:
        rows_axis = "employee"
    metric = (data.get("metric") or "billable").lower()
    if metric not in _VALID_METRICS:
        metric = "billable"

    preset, _created = OrgReportPreset.objects.update_or_create(
        org=org, name=name,
        defaults={
            "rows_axis": rows_axis,
            "metric": metric,
            "expand_all": bool(data.get("expand_all", False)),
            "paginate": bool(data.get("paginate", True)),
            "show_totals": bool(data.get("show_totals", True)),
            "is_default": bool(data.get("is_default", False)),
        },
    )
    # Enforce a single default per org.
    if preset.is_default:
        OrgReportPreset.objects.filter(org=org).exclude(id=preset.id).update(is_default=False)

    return Response({
        "created": _created,
        "id": preset.id,
        "name": preset.name,
        "rows_axis": preset.rows_axis,
        "metric": preset.metric,
        "expand_all": preset.expand_all,
        "paginate": preset.paginate,
        "show_totals": preset.show_totals,
        "is_default": preset.is_default,
    })